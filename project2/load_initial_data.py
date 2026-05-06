"""
load_initial_data.py
Reads initial_data.xlsx and inserts all rows into TransferDB,
bypassing triggers (since the initial dataset intentionally contains
some invalid records that must remain in the database).
"""

import openpyxl
import bcrypt
import mysql.connector
from datetime import datetime, date, time as dtime

from config import db_config

# DB connection
conn = mysql.connector.connect(
    **db_config(),
    allow_local_infile=True
)
cursor = conn.cursor()

# Bypass all triggers for the initial load
cursor.execute("SET @BYPASS_TRIGGERS = 1")


def fmt(v):
    """Return SQL-safe string for a value."""
    if v is None or str(v).strip().upper() == 'NULL':
        return 'NULL'
    if isinstance(v, bool):
        return '1' if v else '0'
    if isinstance(v, (int, float)):
        return str(int(v)) if float(v) == int(float(v)) else str(v)
    if isinstance(v, datetime):
        return f"'{v.strftime('%Y-%m-%d %H:%M:%S')}'"
    if isinstance(v, date):
        return f"'{v.strftime('%Y-%m-%d')}'"
    if isinstance(v, dtime):
        return f"'{v.strftime('%H:%M:%S')}'"
    return f"'{str(v).replace(chr(39), chr(39)+chr(39))}'"


wb = openpyxl.load_workbook('initial_data.xlsx')

# DB Managers
ws = wb['DB Managers']
headers = [c.value for c in ws[1]]
for row in ws.iter_rows(min_row=2, values_only=True):
    if all(v is None for v in row):
        continue
    username, plain_pw = row[0], row[1]
    pw_hash = bcrypt.hashpw(str(plain_pw).encode(), bcrypt.gensalt()).decode()
    sql = (f"INSERT IGNORE INTO DatabaseManager (username, password_hash) "
           f"VALUES ('{username}', '{pw_hash}')")
    cursor.execute(sql)

conn.commit()
print(f"DB Managers loaded: {cursor.rowcount} rows (cumulative)")

# Players
ws = wb['Players']
for row in ws.iter_rows(min_row=2, values_only=True):
    if all(v is None for v in row):
        continue
    pid, uname, plain_pw, name, surname, nat, dob, mv, pos, foot, ht = row
    pw_hash = bcrypt.hashpw(str(plain_pw).encode(), bcrypt.gensalt()).decode()
    dob_str = dob.strftime('%Y-%m-%d') if isinstance(dob, (datetime, date)) else str(dob)
    cursor.execute(
        "INSERT IGNORE INTO Person (person_id, username, password_hash, name, surname, nationality, date_of_birth) "
        "VALUES (%s,%s,%s,%s,%s,%s,%s)",
        (int(pid), uname, pw_hash, name, surname, nat, dob_str)
    )
    cursor.execute(
        "INSERT IGNORE INTO Player (person_id, market_value, main_position, strong_foot, height) "
        "VALUES (%s,%s,%s,%s,%s)",
        (int(pid), float(mv), pos, foot, int(ht))
    )

conn.commit()
print("Players loaded")

# Referees
ws = wb['Referees']
for row in ws.iter_rows(min_row=2, values_only=True):
    if all(v is None for v in row):
        continue
    pid, uname, plain_pw, name, surname, nat, dob, lic, yoe = row
    pw_hash = bcrypt.hashpw(str(plain_pw).encode(), bcrypt.gensalt()).decode()
    dob_str = dob.strftime('%Y-%m-%d') if isinstance(dob, (datetime, date)) else str(dob)
    cursor.execute(
        "INSERT IGNORE INTO Person (person_id, username, password_hash, name, surname, nationality, date_of_birth) "
        "VALUES (%s,%s,%s,%s,%s,%s,%s)",
        (int(pid), uname, pw_hash, name, surname, nat, dob_str)
    )
    cursor.execute(
        "INSERT IGNORE INTO Referee (person_id, license_level, years_of_experience) "
        "VALUES (%s,%s,%s)",
        (int(pid), lic, int(yoe))
    )

conn.commit()
print("Referees loaded")

# Managers
ws = wb['Managers']
for row in ws.iter_rows(min_row=2, values_only=True):
    if all(v is None for v in row):
        continue
    pid, uname, plain_pw, name, surname, nat, dob, formation, exp = row
    pw_hash = bcrypt.hashpw(str(plain_pw).encode(), bcrypt.gensalt()).decode()
    dob_str = dob.strftime('%Y-%m-%d') if isinstance(dob, (datetime, date)) else str(dob)
    cursor.execute(
        "INSERT IGNORE INTO Person (person_id, username, password_hash, name, surname, nationality, date_of_birth) "
        "VALUES (%s,%s,%s,%s,%s,%s,%s)",
        (int(pid), uname, pw_hash, name, surname, nat, dob_str)
    )
    cursor.execute(
        "INSERT IGNORE INTO Manager (person_id, preferred_formation, experience_level) "
        "VALUES (%s,%s,%s)",
        (int(pid), formation, exp)
    )

conn.commit()
print("Managers loaded")

# Stadiums
ws = wb['Stadiums']
for row in ws.iter_rows(min_row=2, values_only=True):
    if all(v is None for v in row):
        continue
    sid, sname, city, cap = row
    cursor.execute(
        "INSERT IGNORE INTO Stadium (stadium_id, stadium_name, city, capacity) "
        "VALUES (%s,%s,%s,%s)",
        (int(sid), sname, city, int(cap))
    )

conn.commit()
print("Stadiums loaded")

# Clubs
ws = wb['Clubs']
for row in ws.iter_rows(min_row=2, values_only=True):
    if all(v is None for v in row):
        continue
    cid, cname, found_year, mgr_id, stad_id = row
    if cid is None or cname is None:
        continue
    mgr_id_val = int(float(str(mgr_id))) if mgr_id and str(mgr_id).upper() != 'NULL' else None
    found_val  = int(found_year) if found_year else None
    stad_val   = int(float(str(stad_id))) if stad_id and str(stad_id).upper() != 'NULL' else None
    cursor.execute(
        "INSERT IGNORE INTO Club (club_id, club_name, foundation_year, manager_id, stadium_id) "
        "VALUES (%s,%s,%s,%s,%s)",
        (int(float(str(cid))), cname, found_val, mgr_id_val, stad_val)
    )

conn.commit()
print("Clubs loaded")

# Competitions
ws = wb['Competitions']
for row in ws.iter_rows(min_row=2, values_only=True):
    if all(v is None for v in row):
        continue
    cid, name, season, country, ctype = row
    cursor.execute(
        "INSERT IGNORE INTO Competition (competition_id, name, season, country, competition_type) "
        "VALUES (%s,%s,%s,%s,%s)",
        (int(cid), name, season, country, ctype)
    )

conn.commit()
print("Competitions loaded")

# Matches
ws = wb['Matches']
for row in ws.iter_rows(min_row=2, values_only=True):
    if all(v is None for v in row):
        continue
    mid, mdate, mtime, stad_id, home_id, away_id, ref_id, comp_id, hg, ag, att = row

    if isinstance(mdate, datetime):
        d = mdate.date()
    else:
        d = mdate

    if isinstance(mtime, dtime):
        t = mtime
    else:
        t = dtime(0, 0)

    mdt = datetime.combine(d, t)

    def nullable_int(v):
        if v is None or str(v).strip().upper() == 'NULL':
            return None
        try:
            return int(float(str(v)))
        except Exception:
            return None

    cursor.execute(
        "INSERT IGNORE INTO `Match` "
        "(match_id, competition_id, home_club_id, away_club_id, stadium_id, referee_id, "
        " match_datetime, home_goals, away_goals, attendance) "
        "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
        (int(mid), int(comp_id), int(home_id), int(away_id), int(stad_id), int(ref_id),
         mdt, nullable_int(hg), nullable_int(ag), nullable_int(att))
    )

conn.commit()
print("Matches loaded")

# Contracts
ws = wb['Contracts']
for row in ws.iter_rows(min_row=2, values_only=True):
    if all(v is None for v in row):
        continue
    cid, pid, club_id, ctype, wage, sdate, edate = row

    def to_date(v):
        if isinstance(v, datetime):
            return v.strftime('%Y-%m-%d')
        if isinstance(v, date):
            return v.strftime('%Y-%m-%d')
        return str(v)

    cursor.execute(
        "INSERT IGNORE INTO Contract (contract_id, player_id, club_id, contract_type, weekly_wage, start_date, end_date) "
        "VALUES (%s,%s,%s,%s,%s,%s,%s)",
        (int(cid), int(pid), int(club_id), ctype, float(wage), to_date(sdate), to_date(edate))
    )

conn.commit()
print("Contracts loaded")

# Transfer Records
ws = wb['Transfer Records']
for row in ws.iter_rows(min_row=2, values_only=True):
    if all(v is None for v in row):
        continue
    tid, pid, from_id, to_id, tdate, fee, ttype = row

    def to_date(v):
        if isinstance(v, datetime):
            return v.strftime('%Y-%m-%d')
        if isinstance(v, date):
            return v.strftime('%Y-%m-%d')
        return str(v)

    from_val = int(float(str(from_id))) if from_id and str(from_id).upper() != 'NULL' else None

    cursor.execute(
        "INSERT IGNORE INTO Transfer_Record "
        "(transfer_id, player_id, from_club_id, to_club_id, transfer_date, transfer_fee, transfer_type) "
        "VALUES (%s,%s,%s,%s,%s,%s,%s)",
        (int(tid), int(pid), from_val, int(to_id), to_date(tdate), float(fee), ttype)
    )

conn.commit()
print("Transfer Records loaded")

# Match Stats
ws = wb['Match Stats']
for row in ws.iter_rows(min_row=2, values_only=True):
    if all(v is None for v in row):
        continue
    mid, pid, club_id, is_starter, mp, pos, goals, assists, yc, rc, rating = row

    def to_bool(v):
        if isinstance(v, bool): return int(v)
        if str(v).strip().upper() in ('TRUE','1','YES'): return 1
        return 0

    cursor.execute(
        "INSERT IGNORE INTO Match_Stats "
        "(match_id, player_id, club_id, is_starter, minutes_played, position_in_match, "
        " goals, assists, yellow_cards, red_cards, rating) "
        "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
        (int(mid), int(pid), int(club_id), to_bool(is_starter),
         int(mp), pos, int(goals), int(assists), int(yc), to_bool(rc), float(rating))
    )

conn.commit()
print("Match Stats loaded")

# Re-enable triggers
cursor.execute("SET @BYPASS_TRIGGERS = 0")
cursor.close()
conn.close()
print("\n✅ Initial data loaded successfully.")
